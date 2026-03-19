import os
from typing import List
from langchain_community.document_loaders import PyPDFLoader, Docx2txtLoader, UnstructuredExcelLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.embeddings import DashScopeEmbeddings
from langchain_milvus import Milvus
from langchain_core.documents import Document
from backend.config import MILVUS_URI, EMBEDDING_MODEL_NAME, QWEN_API_KEY

# 初始化向量化模型（使用 Qwen / DashScope 的 Embedding 服务）
# 注意：需要在环境变量中配置 DASHSCOPE_API_KEY 或 QWEN_API_KEY
embeddings = DashScopeEmbeddings(
    model=EMBEDDING_MODEL_NAME,
    dashscope_api_key=QWEN_API_KEY
)

def _load_excel_fallback(file_path: str) -> List[Document]:
    """
    使用 openpyxl 作为兜底方案读取 Excel，避免因未安装 unstructured 而无法解析。
    将每个工作表拼成一段纯文本，保留表头与行数据（制表符分隔，换行分行）。
    """
    from openpyxl import load_workbook
    docs: List[Document] = []
    wb = load_workbook(filename=file_path, data_only=True)
    for sheet in wb.worksheets:
        lines = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if v is None else str(v) for v in row]
            lines.append("\t".join(values))
        content = f"工作表: {sheet.title}\n" + "\n".join(lines)
        docs.append(Document(page_content=content, metadata={"source": f"{file_path}#{sheet.title}"}))
    return docs

def load_document(file_path: str) -> List[Document]:
    """
    根据文件类型选择合适的解析器加载文档，返回 LangChain 的 Document 列表。
    支持：PDF、DOCX、Excel（.xlsx/.xls）
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        loader = PyPDFLoader(file_path)
    elif ext == ".docx":
        loader = Docx2txtLoader(file_path)
    elif ext in [".xlsx", ".xls"]:
        try:
            loader = UnstructuredExcelLoader(file_path, mode="elements")
            docs = loader.load()
            # 某些环境下 unstructured 可用但产物为空，采用兜底方案
            if not docs:
                return _load_excel_fallback(file_path)
            return docs
        except Exception:
            return _load_excel_fallback(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")
    return loader.load()

def process_documents(file_paths: List[str]):
    """
    处理并入库一批文档：
    1. 加载原始文档
    2. 切分为语义块（chunk）
    3. 生成向量并写入 Milvus 向量数据库
    返回：切分后的文档块数量
    """
    all_docs = []
    for path in file_paths:
        try:
            docs = load_document(path)
            all_docs.extend(docs)
        except Exception as e:
            print(f"Error loading {path}: {e}")

    # 使用递归字符切分器将长文本切为较小片段，便于向量化与召回
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
    splits = text_splitter.split_documents(all_docs)

    # 写入 Milvus（指定集合名为 knowledge_base）
    vector_store = Milvus(
        embedding_function=embeddings,
        connection_args={"uri": MILVUS_URI},
        collection_name="knowledge_base",
        auto_id=True
    )
    
    vector_store.add_documents(splits)
    return len(splits)
